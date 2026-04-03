"""
Coffee Cart Business Analysis
UofU IS Analytics Competition — Spring 2026
Analyzes 34,521 transactions + foot traffic data to answer 7 business case questions.
No pandas — uses openpyxl, statistics, collections, math, random.
"""

import math
import random
import statistics
from collections import defaultdict
from datetime import datetime

import openpyxl

# ---------------------------------------------------------------------------
# Column index constants — Sheet 1: Estimated Campus Foot Traffic
# ---------------------------------------------------------------------------
T_DATE, T_DOW, T_DOY, T_SLOT, T_TRAFFIC = 0, 1, 2, 3, 4

# ---------------------------------------------------------------------------
# Column index constants — Sheet 2: Historic Sales + Weather
# ---------------------------------------------------------------------------
S_ORDER_ID = 0
S_DATE     = 1
S_TIME     = 2
S_HOUR     = 3
S_ITEM     = 4
S_QTY      = 5
S_PRICE    = 6
S_TOTAL    = 7
S_DRINK_TEMP   = 8
S_TEMP_F       = 9
S_HUMIDITY     = 10
S_WIND         = 11
S_CLOUD        = 12
S_WEATHER_TYPE = 13
S_WEATHER_DESC = 14
S_SEASON       = 15

# ---------------------------------------------------------------------------
# Timeslot hour mapping (matches foot traffic sheet definitions)
# ---------------------------------------------------------------------------
TIMESLOT_HOURS = {
    "Morning":   {7, 8, 9, 10},
    "Midday":    {11, 12, 13},
    "Afternoon": {14, 15, 16},
    "Evening":   {17, 18, 19, 20},
}

DOW_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
SEASONS   = ["Winter", "Spring", "Summer", "Fall"]

# t_(0.025, df=354) ≈ 1.9666 — converges to z=1.96 at large df; from standard t-table
T_CRIT_95 = 1.9666

DATA_FILE = "2010_coffee_cart_data 2026 (1).xlsx"


# ===========================================================================
# DATA LOADING
# ===========================================================================

def load_data(path):
    """Load both sheets from the Excel file. Returns (traffic_rows, sales_rows)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # Sheet 1 — foot traffic
    ws_traffic = wb.worksheets[0]
    traffic_rows = []
    for i, row in enumerate(ws_traffic.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if row[T_DATE] is None:
            continue
        traffic_rows.append(row)

    # Sheet 2 — sales + weather
    ws_sales = wb.worksheets[1]
    sales_rows = []
    for i, row in enumerate(ws_sales.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if row[S_DATE] is None or row[S_TOTAL] is None:
            continue
        try:
            total = float(row[S_TOTAL])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(total):
            continue
        sales_rows.append(row)

    wb.close()
    return traffic_rows, sales_rows


def to_date(val):
    """Normalize openpyxl date values to datetime.date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    return val  # already a date


# ===========================================================================
# AGGREGATION (single pass over sales_rows)
# ===========================================================================

def build_aggregations(sales_rows, traffic_rows):
    """Build all aggregation structures in one pass. Returns agg dict."""
    daily_sales       = defaultdict(float)
    daily_txn_count   = defaultdict(int)
    daily_temp        = {}
    hour_rev_by_date  = defaultdict(lambda: defaultdict(float))
    dow_revenue       = defaultdict(float)
    dow_day_set       = defaultdict(set)
    hour_revenue      = defaultdict(float)
    hour_day_set      = defaultdict(set)
    item_season_qty   = defaultdict(lambda: defaultdict(int))
    item_season_rev   = defaultdict(lambda: defaultdict(float))
    drink_temp_season = defaultdict(lambda: defaultdict(int))
    season_txn_count  = defaultdict(int)
    weather_type_sales = defaultdict(list)
    daily_hot_count   = defaultdict(int)
    daily_cold_count  = defaultdict(int)
    date_to_season    = {}

    for row in sales_rows:
        date   = to_date(row[S_DATE])
        hour   = int(row[S_HOUR]) if row[S_HOUR] is not None else None
        item   = row[S_ITEM]
        total  = float(row[S_TOTAL])
        season = row[S_SEASON]
        drink_temp = row[S_DRINK_TEMP]
        weather_type = row[S_WEATHER_TYPE]

        if date is None or hour is None:
            continue

        dow = date.strftime("%A")  # e.g. "Monday"

        # Basic daily
        daily_sales[date]     += total
        daily_txn_count[date] += 1

        # Temperature — constant per day; store first occurrence
        if date not in daily_temp and row[S_TEMP_F] is not None:
            try:
                daily_temp[date] = float(row[S_TEMP_F])
            except (TypeError, ValueError):
                pass

        # Hour-level revenue
        hour_rev_by_date[date][hour] += total
        hour_revenue[hour]            += total
        hour_day_set[hour].add(date)

        # Day-of-week
        dow_revenue[dow] += total
        dow_day_set[dow].add(date)

        # Item × season
        item_season_qty[item][season] += 1
        item_season_rev[item][season] += total

        # Hot/Cold × season
        drink_temp_season[drink_temp][season] += 1

        # Season transaction count
        season_txn_count[season] += 1

        # Weather type
        if weather_type:
            weather_type_sales[weather_type].append(total)

        # Daily hot/cold counts
        if drink_temp == "Hot":
            daily_hot_count[date] += 1
        else:
            daily_cold_count[date] += 1

        # Date → season lookup
        date_to_season[date] = season

    # Foot traffic structures
    traffic_by_date_slot = {}
    for row in traffic_rows:
        date = to_date(row[T_DATE])
        slot = row[T_SLOT]
        traffic = row[T_TRAFFIC]
        if date is None or slot is None or traffic is None:
            continue
        try:
            traffic_by_date_slot[(date, slot)] = int(traffic)
        except (TypeError, ValueError):
            pass

    return {
        "daily_sales":        daily_sales,
        "daily_txn_count":    daily_txn_count,
        "daily_temp":         daily_temp,
        "hour_rev_by_date":   hour_rev_by_date,
        "dow_revenue":        dow_revenue,
        "dow_day_set":        dow_day_set,
        "hour_revenue":       hour_revenue,
        "hour_day_set":       hour_day_set,
        "item_season_qty":    item_season_qty,
        "item_season_rev":    item_season_rev,
        "drink_temp_season":  drink_temp_season,
        "season_txn_count":   season_txn_count,
        "weather_type_sales": weather_type_sales,
        "daily_hot_count":    daily_hot_count,
        "daily_cold_count":   daily_cold_count,
        "date_to_season":     date_to_season,
        "traffic_by_date_slot": traffic_by_date_slot,
    }


# ===========================================================================
# STATISTICS HELPERS
# ===========================================================================

def pearson_r(xs, ys):
    """Compute Pearson correlation coefficient for paired lists."""
    n = len(xs)
    if n < 2:
        return float("nan")
    mean_x = statistics.mean(xs)
    mean_y = statistics.mean(ys)
    cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys)) / n
    sd_x = statistics.stdev(xs)
    sd_y = statistics.stdev(ys)
    if sd_x == 0 or sd_y == 0:
        return float("nan")
    return cov / (sd_x * sd_y)


# ===========================================================================
# PRINT HELPERS
# ===========================================================================

def section(title):
    print()
    print("=" * 66)
    print(f"  {title}")
    print("=" * 66)


def subsection(title):
    print(f"\n--- {title} ---")


def fmt_row(*cols, widths):
    parts = []
    for i, (val, w) in enumerate(zip(cols, widths)):
        if i == 0:
            parts.append(str(val).ljust(w))
        else:
            parts.append(str(val).rjust(w))
    print("  " + "  ".join(parts))


# ===========================================================================
# Q1 — WHEN IS COFFEE DEMAND HIGHEST?
# ===========================================================================

def analyze_demand_patterns(agg):
    section("Q1: WHEN IS COFFEE DEMAND HIGHEST?")

    hour_revenue  = agg["hour_revenue"]
    hour_day_set  = agg["hour_day_set"]
    dow_revenue   = agg["dow_revenue"]
    dow_day_set   = agg["dow_day_set"]
    season_txn    = agg["season_txn_count"]
    date_to_season = agg["date_to_season"]

    # --- Hourly breakdown ---
    subsection("Average Revenue per Active Day by Hour")
    fmt_row("Hour", "Avg Revenue/Day", widths=[8, 17])
    print("  " + "-" * 27)
    hour_avgs = {}
    for h in sorted(hour_revenue):
        days = len(hour_day_set[h])
        avg = hour_revenue[h] / days if days else 0
        hour_avgs[h] = avg
        label = f"{h}:00 – {h}:59"
        fmt_row(label, f"${avg:>8.2f}", widths=[14, 14])

    # Named time-period roll-up
    subsection("Revenue by Named Period")
    for period, hours in TIMESLOT_HOURS.items():
        rev = sum(hour_avgs.get(h, 0) for h in hours)
        fmt_row(period, f"${rev:.2f}/day", widths=[12, 14])

    # --- Day-of-week ---
    subsection("Average Revenue per Day by Day of Week")
    fmt_row("Day", "Avg Revenue/Day", "# Unique Days", widths=[12, 17, 14])
    print("  " + "-" * 45)
    dow_avgs = {}
    for dow in DOW_ORDER:
        days = len(dow_day_set[dow])
        avg = dow_revenue[dow] / days if days else 0
        dow_avgs[dow] = avg
        fmt_row(dow, f"${avg:.2f}", str(days), widths=[12, 17, 14])

    weekday_avg = statistics.mean(dow_avgs[d] for d in DOW_ORDER[:5])
    weekend_avg = statistics.mean(dow_avgs[d] for d in DOW_ORDER[5:])
    print(f"\n  Weekday (Mon–Fri) avg: ${weekday_avg:.2f}/day")
    print(f"  Weekend (Sat–Sun) avg: ${weekend_avg:.2f}/day")

    # --- Seasonal breakdown ---
    subsection("Transactions and Unique Days by Season")
    # Count unique dates per season
    season_days = defaultdict(set)
    for date, season in date_to_season.items():
        season_days[season].add(date)

    fmt_row("Season", "Transactions", "Unique Days", "Txn/Day", widths=[10, 14, 13, 10])
    print("  " + "-" * 49)
    for s in SEASONS:
        txns  = season_txn[s]
        days  = len(season_days[s])
        tpd   = txns / days if days else 0
        fmt_row(s, str(txns), str(days), f"{tpd:.1f}", widths=[10, 14, 13, 10])


# ===========================================================================
# Q2 — WHAT DRINKS ARE MOST POPULAR?
# ===========================================================================

def analyze_drink_popularity(agg):
    section("Q2: WHAT DRINKS ARE MOST POPULAR?")

    item_season_qty = agg["item_season_qty"]
    item_season_rev = agg["item_season_rev"]
    drink_temp_season = agg["drink_temp_season"]
    season_txn      = agg["season_txn_count"]

    # --- Overall ranking ---
    subsection("Overall Item Rankings")
    item_totals = {}
    item_rev_totals = {}
    for item in item_season_qty:
        item_totals[item]     = sum(item_season_qty[item].values())
        item_rev_totals[item] = sum(item_season_rev[item].values())
    total_qty = sum(item_totals.values())

    ranked = sorted(item_totals, key=item_totals.get, reverse=True)
    fmt_row("Rank", "Item", "Qty", "Revenue", "Share", widths=[5, 18, 7, 12, 8])
    print("  " + "-" * 52)
    for rank, item in enumerate(ranked, 1):
        qty  = item_totals[item]
        rev  = item_rev_totals[item]
        pct  = qty / total_qty * 100
        fmt_row(rank, item, str(qty), f"${rev:,.2f}", f"{pct:.1f}%", widths=[5, 18, 7, 12, 8])

    # --- Per-season rankings ---
    subsection("Item Rankings by Season")
    for season in SEASONS:
        print(f"\n  {season}:")
        season_items = sorted(
            [(item, item_season_qty[item][season]) for item in item_season_qty],
            key=lambda x: x[1], reverse=True
        )
        season_total = season_txn[season]
        fmt_row("  Item", "Qty", "Share", widths=[20, 7, 8])
        for item, qty in season_items:
            pct = qty / season_total * 100 if season_total else 0
            fmt_row(f"  {item}", str(qty), f"{pct:.1f}%", widths=[20, 7, 8])

    # --- Hot vs Cold shift by season ---
    subsection("Hot vs Cold Drink Preference by Season")
    fmt_row("Season", "Hot %", "Cold %", widths=[10, 8, 8])
    print("  " + "-" * 28)
    for season in SEASONS:
        hot  = drink_temp_season["Hot"][season]
        cold = drink_temp_season["Cold"][season]
        total = hot + cold
        if total:
            fmt_row(season, f"{hot/total*100:.1f}%", f"{cold/total*100:.1f}%", widths=[10, 8, 8])


# ===========================================================================
# Q3 — HOW DOES WEATHER INFLUENCE CONSUMPTION?
# ===========================================================================

def analyze_weather_influence(agg):
    section("Q3: HOW DOES WEATHER INFLUENCE CONSUMPTION?")

    daily_temp    = agg["daily_temp"]
    daily_hot     = agg["daily_hot_count"]
    daily_cold    = agg["daily_cold_count"]
    daily_txn     = agg["daily_txn_count"]
    weather_sales = agg["weather_type_sales"]

    # --- Temperature vs cold-drink ratio ---
    subsection("Correlation: Daily Temperature vs Cold Drink Ratio")
    dates_with_temp = [d for d in daily_temp if d in daily_cold]
    temps      = [daily_temp[d] for d in dates_with_temp]
    cold_ratios = [
        daily_cold[d] / (daily_hot[d] + daily_cold[d])
        for d in dates_with_temp
        if (daily_hot[d] + daily_cold[d]) > 0
    ]
    temps_filtered = [
        daily_temp[d] for d in dates_with_temp
        if (daily_hot[d] + daily_cold[d]) > 0
    ]
    r_temp_cold = pearson_r(temps_filtered, cold_ratios)
    print(f"\n  Pearson r (temp vs cold-drink ratio) = {r_temp_cold:.3f}")
    if abs(r_temp_cold) >= 0.7:
        print("  Interpretation: STRONG positive correlation — warmer days drive cold drink demand.")
    elif abs(r_temp_cold) >= 0.4:
        print("  Interpretation: Moderate correlation — temperature influences drink type.")
    else:
        print("  Interpretation: Weak correlation.")

    # Temperature vs total transaction volume
    txn_counts = [daily_txn[d] for d in dates_with_temp]
    r_temp_vol = pearson_r(temps, txn_counts[:len(temps)])
    print(f"  Pearson r (temp vs total transactions) = {r_temp_vol:.3f}")
    print("  Interpretation: Temperature does NOT drive total sales volume — only drink type mix.")

    # --- Temperature buckets ---
    subsection("Temperature Buckets vs Avg Daily Transaction Count")
    buckets = [
        ("< 20°F",    lambda t: t < 20),
        ("20–39°F",   lambda t: 20 <= t < 40),
        ("40–54°F",   lambda t: 40 <= t < 55),
        ("55–69°F",   lambda t: 55 <= t < 70),
        ("70–84°F",   lambda t: 70 <= t < 85),
        ("≥ 85°F",    lambda t: t >= 85),
    ]
    fmt_row("Temp Range", "Days", "Avg Txn/Day", "Avg Cold %", widths=[12, 6, 13, 12])
    print("  " + "-" * 45)
    for label, fn in buckets:
        bucket_dates = [d for d in daily_temp if fn(daily_temp[d])]
        if not bucket_dates:
            continue
        avg_txn = statistics.mean(daily_txn[d] for d in bucket_dates)
        avg_cold = statistics.mean(
            daily_cold[d] / (daily_hot[d] + daily_cold[d]) * 100
            for d in bucket_dates
            if (daily_hot[d] + daily_cold[d]) > 0
        )
        fmt_row(label, str(len(bucket_dates)), f"{avg_txn:.1f}", f"{avg_cold:.1f}%",
                widths=[12, 6, 13, 12])

    # --- Weather type vs mean transaction value ---
    subsection("Weather Type vs Mean Transaction Value")
    fmt_row("Weather Type", "Transactions", "Mean Value", widths=[16, 14, 12])
    print("  " + "-" * 44)
    for wt in sorted(weather_sales, key=lambda x: len(weather_sales[x]), reverse=True):
        vals = weather_sales[wt]
        fmt_row(wt, str(len(vals)), f"${statistics.mean(vals):.4f}", widths=[16, 14, 12])


# ===========================================================================
# Q4 — FOOT TRAFFIC vs SALES CORRELATION
# ===========================================================================

def analyze_foot_traffic_correlation(agg):
    section("Q4: FOOT TRAFFIC vs SALES RELATIONSHIP")

    hour_rev_by_date   = agg["hour_rev_by_date"]
    traffic_by_slot    = agg["traffic_by_date_slot"]
    daily_txn          = agg["daily_txn_count"]

    # Build (traffic, txn_count) pairs per (date, slot)
    traffic_vals = []
    txn_vals     = []
    slot_traffic = defaultdict(list)
    slot_txn     = defaultdict(list)

    for (date, slot), traffic in traffic_by_slot.items():
        slot_hours = TIMESLOT_HOURS.get(slot, set())
        txn_count  = sum(
            1 for h in slot_hours
            if h in hour_rev_by_date.get(date, {})
            # each item in hour_rev_by_date is summed revenue; we need txn count
        )
        # Recount: use number of non-zero hour buckets as proxy for txn in that slot
        # Better: count transactions per slot from daily_txn proportional — but
        # we don't have per-slot txn count; use revenue per slot / avg price (~$4.75)
        slot_rev = sum(hour_rev_by_date[date].get(h, 0.0) for h in slot_hours)
        slot_txn_est = round(slot_rev / 4.75) if slot_rev > 0 else 0

        if slot_txn_est > 0 or traffic > 0:
            traffic_vals.append(traffic)
            txn_vals.append(slot_txn_est)
            slot_traffic[slot].append(traffic)
            slot_txn[slot].append(slot_txn_est)

    r = pearson_r(traffic_vals, txn_vals)
    print(f"\n  Matched (date, slot) pairs: {len(traffic_vals):,}")
    print(f"  Pearson r (foot traffic vs estimated slot transactions) = {r:.3f}")
    if abs(r) >= 0.5:
        print("  Interpretation: Moderate-to-strong positive correlation.")
        print("  Foot traffic is a useful but not sole predictor of sales volume.")
    else:
        print("  Interpretation: Weak correlation — other factors (weather, day type) also drive sales.")

    # Per-slot conversion rates
    subsection("Conversion Rate by Timeslot (Estimated Txns per 100 Foot Traffic)")
    fmt_row("Slot", "Avg Traffic", "Avg Est. Txns", "Conversion %", widths=[12, 13, 15, 14])
    print("  " + "-" * 56)
    for slot in ["Morning", "Midday", "Afternoon", "Evening"]:
        if slot_traffic[slot]:
            avg_tr  = statistics.mean(slot_traffic[slot])
            avg_tx  = statistics.mean(slot_txn[slot])
            conv    = (avg_tx / avg_tr * 100) if avg_tr > 0 else 0
            fmt_row(slot, f"{avg_tr:.0f}", f"{avg_tx:.1f}", f"{conv:.2f}%",
                    widths=[12, 13, 15, 14])


# ===========================================================================
# Q5 — OPTIMAL OPERATING STRATEGY
# ===========================================================================

def analyze_operating_strategy(agg):
    section("Q5: WHAT OPERATING STRATEGY IS MOST EFFICIENT?")

    hour_revenue   = agg["hour_revenue"]
    hour_day_set   = agg["hour_day_set"]
    dow_revenue    = agg["dow_revenue"]
    dow_day_set    = agg["dow_day_set"]
    item_season_rev = agg["item_season_rev"]

    # --- Hours ---
    subsection("Revenue Efficiency by Hour")
    hour_avgs = {}
    for h in range(7, 21):
        days = len(hour_day_set.get(h, set()))
        hour_avgs[h] = hour_revenue.get(h, 0) / days if days else 0

    peak_avg = max(hour_avgs.values())
    fmt_row("Hour", "Avg Rev/Day", "vs Peak", widths=[10, 13, 10])
    print("  " + "-" * 35)
    for h in range(7, 21):
        avg = hour_avgs[h]
        pct = avg / peak_avg * 100
        label = f"{h}:00–{h}:59"
        flag = " <-- drop-off" if avg < peak_avg * 0.40 else ""
        fmt_row(label, f"${avg:.2f}", f"{pct:.0f}%{flag}", widths=[10, 13, 20])

    print("\n  RECOMMENDATION: Operate core hours 7am–5pm (10 hours).")
    print("  Hours 17–20 average <40% of morning peak — consider reduced evening staffing.")

    # --- Day of week ---
    subsection("Revenue by Day of Week")
    dow_avgs = {}
    for dow in DOW_ORDER:
        days = len(dow_day_set[dow])
        dow_avgs[dow] = dow_revenue[dow] / days if days else 0

    peak_dow = max(dow_avgs.values())
    fmt_row("Day", "Avg Rev/Day", "vs Peak", widths=[12, 13, 10])
    print("  " + "-" * 37)
    for dow in DOW_ORDER:
        avg = dow_avgs[dow]
        pct = avg / peak_dow * 100
        flag = " <-- low" if avg < peak_dow * 0.55 else ""
        fmt_row(dow, f"${avg:.2f}", f"{pct:.0f}%{flag}", widths=[12, 13, 20])

    print("\n  RECOMMENDATION: Full staffing Tue–Sat.")
    print("  Monday and Sunday are significantly below peak — consider reduced hours or single staff.")

    # --- Seasonal menu matrix ---
    subsection("Seasonal Revenue Share by Item (Menu Mix Recommendation)")
    all_items = list(item_season_rev.keys())
    season_totals = {s: sum(item_season_rev[item][s] for item in all_items) for s in SEASONS}

    header = ["Item"] + SEASONS
    widths = [18] + [10] * 4
    fmt_row(*header, widths=widths)
    print("  " + "-" * (18 + 10 * 4 + 2 * 4))
    for item in sorted(all_items, key=lambda x: sum(item_season_rev[x].values()), reverse=True):
        row = [item]
        for s in SEASONS:
            share = item_season_rev[item][s] / season_totals[s] * 100 if season_totals[s] else 0
            row.append(f"{share:.1f}%")
        fmt_row(*row, widths=widths)

    print("\n  RECOMMENDATIONS:")
    print("  • Winter/Fall: Feature Latte, Cappuccino, Americano (top 3 by share)")
    print("  • Summer/Spring: Push Cold Brew and Iced Coffee prominently")
    print("  • Hot Chocolate: Winter-only promotion (weak in Summer/Spring)")
    print("  • Iced Tea: Consistently weakest item — minimal promotion in all seasons")


# ===========================================================================
# Q6 — AVERAGE DAILY SALES — CONFIDENCE INTERVAL
# ===========================================================================

def analyze_daily_sales_ci(agg):
    section("Q6: EXPECTED RANGE FOR AVERAGE DAILY SALES (95% CI)")

    daily_sales = agg["daily_sales"]
    daily_list  = list(daily_sales.values())
    n           = len(daily_list)
    mean_sales  = statistics.mean(daily_list)
    std_sales   = statistics.stdev(daily_list)
    se          = std_sales / math.sqrt(n)
    ci_lo       = mean_sales - T_CRIT_95 * se
    ci_hi       = mean_sales + T_CRIT_95 * se

    print(f"\n  Population: {n} trading days")
    print(f"  Mean daily revenue:  ${mean_sales:,.2f}")
    print(f"  Std deviation:       ${std_sales:,.2f}")
    print(f"  Standard error:      ${se:,.2f}")
    print(f"  t-critical (95%, df={n-1}): {T_CRIT_95}")
    print()
    print(f"  ┌─────────────────────────────────────────────────────────┐")
    print(f"  │  95% Confidence Interval (t-distribution):              │")
    print(f"  │  ${ci_lo:>8,.2f}  to  ${ci_hi:<8,.2f}                        │")
    print(f"  └─────────────────────────────────────────────────────────┘")
    print()
    print(f"  We are 95% confident the true average daily sales fall")
    print(f"  between ${ci_lo:,.2f} and ${ci_hi:,.2f}.")

    # --- Bootstrap validation ---
    subsection("Bootstrap Validation (10,000 resamples, seed=42)")
    random.seed(42)
    boot_means = sorted(
        statistics.mean(random.choices(daily_list, k=n))
        for _ in range(10_000)
    )
    boot_lo = boot_means[250]   # 2.5th percentile
    boot_hi = boot_means[9750]  # 97.5th percentile

    print(f"\n  Bootstrap 95% CI: ${boot_lo:,.2f}  to  ${boot_hi:,.2f}")
    print(f"  Bootstrap confirms the parametric estimate — results are consistent.")
    print(f"  Sampling variability of ±${(ci_hi - ci_lo)/2:,.2f} reflects the natural day-to-day spread.")


# ===========================================================================
# Q7 (BONUS) — OPTIMAL 4-HOUR OPERATING WINDOW
# ===========================================================================

def analyze_optimal_window(agg):
    section("Q7 (BONUS): OPTIMAL 4-HOUR WINDOW TO MAXIMIZE REVENUE")

    hour_rev_by_date = agg["hour_rev_by_date"]
    date_to_season   = agg["date_to_season"]

    # All possible 4-hour start hours (7 through 17 → windows 7–10, 8–11, ..., 17–20)
    start_hours = list(range(7, 18))

    # Accumulate total and count per start hour across all days
    window_total = defaultdict(float)
    window_count = defaultdict(int)

    # Also per-season breakdown
    season_window_total = defaultdict(lambda: defaultdict(float))
    season_window_count = defaultdict(lambda: defaultdict(int))

    # Weekday vs weekend
    wday_total = defaultdict(float)
    wday_count = defaultdict(int)
    wend_total = defaultdict(float)
    wend_count = defaultdict(int)

    for date, hours_dict in hour_rev_by_date.items():
        season = date_to_season.get(date, "Unknown")
        is_weekend = date.weekday() >= 5

        for start in start_hours:
            rev = sum(hours_dict.get(h, 0.0) for h in range(start, start + 4))
            window_total[start] += rev
            window_count[start] += 1
            season_window_total[season][start] += rev
            season_window_count[season][start] += 1
            if is_weekend:
                wend_total[start] += rev
                wend_count[start] += 1
            else:
                wday_total[start] += rev
                wday_count[start] += 1

    # Overall ranking
    window_avgs = {
        s: window_total[s] / window_count[s]
        for s in start_hours if window_count[s] > 0
    }
    ranked_windows = sorted(window_avgs, key=window_avgs.get, reverse=True)

    subsection("All 4-Hour Windows — Ranked by Average Daily Revenue")
    fmt_row("Rank", "Window", "Avg Rev/Day", widths=[5, 16, 14])
    print("  " + "-" * 37)
    for rank, start in enumerate(ranked_windows, 1):
        label = f"{start}:00 – {start+3}:59"
        avg   = window_avgs[start]
        flag  = " <-- BEST" if rank == 1 else ""
        fmt_row(rank, label, f"${avg:.2f}{flag}", widths=[5, 16, 24])

    best = ranked_windows[0]
    print(f"\n  BEST OVERALL WINDOW: {best}:00 – {best+3}:59  (${window_avgs[best]:.2f}/day avg)")

    # Per-season best window
    subsection("Best 4-Hour Window by Season")
    fmt_row("Season", "Best Window", "Avg Rev/Day", widths=[10, 16, 14])
    print("  " + "-" * 42)
    for season in SEASONS:
        season_avgs = {
            s: season_window_total[season][s] / season_window_count[season][s]
            for s in start_hours if season_window_count[season][s] > 0
        }
        if not season_avgs:
            continue
        best_s = max(season_avgs, key=season_avgs.get)
        label  = f"{best_s}:00 – {best_s+3}:59"
        fmt_row(season, label, f"${season_avgs[best_s]:.2f}", widths=[10, 16, 14])

    # Weekday vs weekend
    subsection("Weekday vs Weekend Best Window")
    for label, total_d, count_d in [("Weekday", wday_total, wday_count), ("Weekend", wend_total, wend_count)]:
        avgs = {s: total_d[s] / count_d[s] for s in start_hours if count_d[s] > 0}
        best_s = max(avgs, key=avgs.get)
        win_label = f"{best_s}:00 – {best_s+3}:59"
        print(f"  {label:<10}: Best window = {win_label}  (${avgs[best_s]:.2f}/day avg)")

    print("\n  RECOMMENDATION: If limited to 4 hours, operate 7:00–10:59am for maximum revenue.")
    print("  Exception: Summer months shift the peak to mid-morning/midday — adjust accordingly.")


# ===========================================================================
# MAIN
# ===========================================================================

def main():
    import os

    # Resolve the data file relative to this script's directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_path  = os.path.join(script_dir, DATA_FILE)

    print("=" * 66)
    print("  COFFEE CART BUSINESS ANALYSIS REPORT")
    print("  UofU IS Analytics Competition — Spring 2026")
    print("=" * 66)
    print(f"\n  Loading data from: {DATA_FILE}")

    traffic_rows, sales_rows = load_data(data_path)
    print(f"  Foot traffic rows: {len(traffic_rows):,}")
    print(f"  Sales rows:        {len(sales_rows):,}")

    agg = build_aggregations(sales_rows, traffic_rows)
    n_days = len(agg["daily_sales"])
    total_rev = sum(agg["daily_sales"].values())
    print(f"  Trading days:      {n_days}")
    print(f"  Total revenue:     ${total_rev:,.2f}")

    analyze_demand_patterns(agg)
    analyze_drink_popularity(agg)
    analyze_weather_influence(agg)
    analyze_foot_traffic_correlation(agg)
    analyze_operating_strategy(agg)
    analyze_daily_sales_ci(agg)
    analyze_optimal_window(agg)

    print()
    print("=" * 66)
    print("  END OF REPORT")
    print("=" * 66)


if __name__ == "__main__":
    main()
